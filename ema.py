import pandas as pd
import numpy as np
import pytz
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# For loading current price data
price_input = '' # add file name containing price data here
df = pd.read_excel(price_input)
df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d %H:%M:%S %Z', utc=True)
df = df.set_index('Date')

df['Price_Change'] = df['Price'].pct_change() # Creates a new column 'Price_Change' by calculating the percentage change based on values from 'Price' column
daily_std_dev = df['Price_Change'].std() # Calculates std dev

# For calculating Exponential Moving Average (EMA)

def get_ema(data, span):
    return data['Price'].ewm(span=span, adjust=False).mean # Selects values from 'Price' column to calculate the mean of the exponential weighted window, thus computing the EMA

df['EMA'] = get_ema(df, 30) # Calculates EMA over a 30 day period and updates dataframe

# For setting the initial conditions required to forecast prices
def forecast_prices(df, end_date, daily_std_dev):
    last_date = df.index[-1] # Selects the last row in dataframe
    last_ema = df['EMA'].iloc[-1] # Selects the last EMA (most recent) 
    last_price = df['Price'].iloc[-1] # Selects the last price (most recent)

    future_dates = pd.date_range(start=last_date + timedelta(days=1), end=end_date, freq='D', tz='UTC') # For creating a range of dates from the day after the last historical date up to the specified end date
    
    # For creating random variations based on the daily standard deviation and calculates cumulative effects
    variations = np.random.normal(0, daily_std_dev, len(future_dates))
    cumulative_variations = np.cumprod(1 + variations)

    # For calculating future prices based on the last EMA adjusted by the cumulative variations
    future_prices = last_ema * cumulative_variations

    # For updating dateframe with forecasted projections
    future_df = pd.DataFrame({
        'Date': future_dates,
        'Price': future_prices,
        'EMA': [last_ema] * len(future_dates)
    })
    future_df = future_df.set_index('Date')

    # For calculating the Price_Change in forecasted projections
    future_df['Price_Change'] = future_df['Price'].pct_change() # Selects forecasted prices and calculates percentage change
    future_df.loc[future_df.index[0], 'Price_Change'] = (future_df['Price'].iloc[0] - last_price) / last_price

    return future_df

# For establishing a target date to end forecasted projections
target_date = datetime(2025, 9, 6, tzinfo=pytz.UTC) # Target date is Sept-6-2025, based on miner depreciation
forecast = forecast_prices(df, target_date, daily_std_dev)

# For unifying current data and forecasted projections
result = pd.concat([df, forecast])

# For ensuring the date format matches that of the input file
result.index = result.index.strftime('%Y-%m-%d %H:%M:%S UTC')
result = result.reset_index()

# For saving dataframe to an xlsx file
output_file = '' # add desired file output name here
result.to_excel(output_file, index=False, engine='openpyxl')

# The following section simply applies to the formatting of the xlxs file
wb = load_workbook(output_file)
ws = wb.active

# For date and price styling
date_style = NamedStyle(name="datetime", number_format='YYYY-MM-DD HH:MM:SS UTC')
price_style = NamedStyle(name="currency", number_format='$#,##0.00')

for row in ws['A2:A'+str(ws.max_row)]:  # Skip header
    cell = row[0]
    cell.style = date_style

for row in ws['B2:C'+str(ws.max_row)]:  # Skip header
    for cell in row:
        cell.style = price_style

# Save the updated xlsx file with proper formatting
wb.save(output_file)

# Print stametemnts added for my liking, so that I can see on my terminal the script finalizing
print(f"Forecast generated up to {target_date.strftime('%Y-%m-%d %H:%M:%S %Z')}")
print(f"Results saved to '{output_file}'")
print(f"Daily standard deviation of price changes: {daily_std_dev:.4f}")
